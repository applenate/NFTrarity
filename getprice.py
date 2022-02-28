#coding:UTF-8
import os
from requests_html import HTMLSession
import json
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import datetime

site_url = "https://opensea.io/assets/"
token_contract = "0x534d37c630b7e4d2a6c1e064f3a2632739e9ee04"
#填写token id 区间,冷兔总量7502，id是0~7501
id_start = 0
id_end = 7501
# num_tokens = 7502
write_fileName = "nfts_冷兔_price.xlsx" #导出文件名
read_filename = "" #读NFT属性、稀缺排名Excel,后续做筛查

#完整url https://opensea.io/assets/0x534d37c630b7e4d2a6c1e064f3a2632739e9ee04/207

session = HTMLSession()
seprate ='__________________________________________________________________________\n'

print(seprate)
print(time.strftime("    开始读取opensea价格  |  服务器时间 %Y-%m-%d %H:%M:%S", time.localtime())) 
print("    合约:%s"%token_contract)
print("    开始id:%d"%id_start)
print("    结束id:%d"%id_end)
print("    导出:%s"%write_fileName)

#创建文件
def create(write_fileName):
	wb = Workbook()
	ws = wb.active
	ws.title = "Price" #sheet 名称
	content = ['Link','ID','Price','Type_text','Time'] #标题
	ws.append(content)
	wb.save(write_fileName)

#写入行数据，无论该行有多少列数据，第一行row = 1，是表头
def insert(row,content,link):
	print(content)
	wb = load_workbook(write_fileName)
	ws = wb["Price"]
	for col in range(1,len(content)+2):
		if col == 1:
			ws.cell(column = col,row = row ,value = 'opensea')
			ws.cell(column = col,row = row ).hyperlink = link
			ws.cell(column = col,row = row ).style = 'Hyperlink'
		else:
			ws.cell(column = col,row = row ,value = content[col-2])

	wb.save(write_fileName)


def getNFT_price(tokenid,token_contract,jindu):
	url = site_url+token_contract+"/"+str(tokenid)
	r = session.get(url).html
	content_price = r.find("div.Overflowreact__OverflowContainer-sc-7qr9y8-0.jPSCbX.Price--amount")
	
	#print(r.find("div.TradeStation--ask-label"))
	content_type = r.find("div.TradeStation--ask-label")
	if len(content_type) != 0 :
		price_type = r.find("div.TradeStation--ask-label")[0].text
	else :
		price_type = "<none>"

	if len(content_price) != 0 :
		price = content_price[0].text #价格为str
	else :
		price = "<none>"
	print("<%.2f%%>  id: %d  ,  %s ETH  ,  %s"%(jindu,tokenid,price,price_type))

	price_info = ['#'+str(tokenid),price,price_type,datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
	insert(tokenid+2,price_info,url)  #写入Excel row = tokenid + 2


print(seprate)
time_start = time.time() #开始时间

count = 0 #计数

#如果没有文件，先创建再插数据；如果有文件，直接插数据
if os.path.exists(write_fileName) is not True :
	create(write_fileName)

num_tokens = (id_end - id_start) + 1

for i in range(id_start,id_end + 1):  #num_tokens
	jindu = ((count+1)/num_tokens)  #float 显示进度
	price = getNFT_price(i,token_contract,jindu * 100) #抓取并写Excel，指定tokenid
	count += 1
	
	#防反爬虫机制，每隔50个，休息1秒
	if (i+1)%50 == 0:
		time_now = time.time() 
		time_used = time_now - time_start
		time_remain = ( time_used / jindu ) * (1 - jindu) + 1
		print("-  休息1秒 （ 当前Id = %d）  预计还需 "%i,datetime.timedelta(seconds=int(time_remain)),"   已用时 ",datetime.timedelta(seconds=int(time_used)))
		time.sleep(1)

time_end = time.time() #结束时间
deltatime = int(time_end - time_start)
print("    总耗时： ",datetime.timedelta(seconds=deltatime))
print(seprate)
